"""
Translation Manager - Unified service for all translation types
Manages 3-tier subscription system:
- FREE: Google Translate (via deep-translator)
- DEEPL: DeepL API
- AI: AI-powered translation (ChatGPT, Claude, etc.)
"""

import os
import sys
from pathlib import Path
from typing import Optional, Dict
from enum import Enum

# Add parent directories to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

# Import document translator
from document_translator import (
    DocumentTranslator,
    GoogleTranslateService,
    DeepLTranslateService
)

# Import AI translator
from ai_translator_addon import (
    ChatGPTTranslationService,
    ClaudeTranslationService,
    GeminiTranslationService
)

# Import image translator components
sys.path.insert(0, str(Path(__file__).parent.parent.parent / 'image-translator' / 'src' / 'image-translator'))
try:
    from core.pipeline import TranslationPipeline
    IMAGE_TRANSLATOR_AVAILABLE = True
except ImportError:
    IMAGE_TRANSLATOR_AVAILABLE = False
    print("Warning: Image translator not available")


class UserTier(Enum):
    """User subscription tiers"""
    FREE = "free"
    DEEPL = "deepl"
    AI = "ai"


class TranslationManager:
    """
    Unified translation manager for text, documents, and images
    """

    def __init__(self):
        """Initialize translation manager"""
        self.config = self._load_config()
        self.image_pipeline_cache = {}

    def _load_config(self) -> Dict:
        """Load configuration from environment or defaults"""
        return {
            'deepl_api_key': os.environ.get('DEEPL_API_KEY'),
            'openai_api_key': os.environ.get('OPENAI_API_KEY'),
            'anthropic_api_key': os.environ.get('ANTHROPIC_API_KEY'),
            'google_api_key': os.environ.get('GOOGLE_API_KEY'),
            'ai_default_model': os.environ.get('AI_DEFAULT_MODEL', 'gpt-4o-mini'),
            'ai_service': os.environ.get('AI_SERVICE', 'chatgpt')  # chatgpt, claude, gemini
        }

    def _get_translation_service(self, tier: UserTier, source_lang: str, target_lang: str):
        """
        Get appropriate translation service based on user tier

        Args:
            tier: User subscription tier
            source_lang: Source language code
            target_lang: Target language code

        Returns:
            Translation service instance
        """
        if tier == UserTier.FREE:
            # Free tier: Google Translate
            return GoogleTranslateService(
                source_lang=source_lang,
                target_lang=target_lang
            )

        elif tier == UserTier.DEEPL:
            # DeepL tier: DeepL API
            api_key = self.config.get('deepl_api_key')
            if not api_key:
                raise ValueError("DeepL API key not configured. Please set DEEPL_API_KEY environment variable.")

            return DeepLTranslateService(
                source_lang=source_lang,
                target_lang=target_lang,
                api_key=api_key
            )

        elif tier == UserTier.AI:
            # AI tier: ChatGPT, Claude, or Gemini
            ai_service = self.config.get('ai_service', 'chatgpt')

            if ai_service == 'chatgpt':
                api_key = self.config.get('openai_api_key')
                if not api_key:
                    raise ValueError("OpenAI API key not configured. Please set OPENAI_API_KEY environment variable.")

                return ChatGPTTranslationService(
                    source_lang=source_lang,
                    target_lang=target_lang,
                    api_key=api_key,
                    model=self.config.get('ai_default_model', 'gpt-4o-mini')
                )

            elif ai_service == 'claude':
                api_key = self.config.get('anthropic_api_key')
                if not api_key:
                    raise ValueError("Anthropic API key not configured. Please set ANTHROPIC_API_KEY environment variable.")

                return ClaudeTranslationService(
                    source_lang=source_lang,
                    target_lang=target_lang,
                    api_key=api_key,
                    model='claude-3-5-haiku-20241022'
                )

            elif ai_service == 'gemini':
                api_key = self.config.get('google_api_key')
                if not api_key:
                    raise ValueError("Google API key not configured. Please set GOOGLE_API_KEY environment variable.")

                return GeminiTranslationService(
                    source_lang=source_lang,
                    target_lang=target_lang,
                    api_key=api_key,
                    model='gemini-1.5-flash'
                )

            else:
                raise ValueError(f"Unknown AI service: {ai_service}")

        else:
            raise ValueError(f"Unknown tier: {tier}")

    def translate_text(
        self,
        text: str,
        source_lang: str,
        target_lang: str,
        tier: UserTier
    ) -> str:
        """
        Translate plain text

        Args:
            text: Text to translate
            source_lang: Source language code
            target_lang: Target language code
            tier: User subscription tier

        Returns:
            Translated text
        """
        service = self._get_translation_service(tier, source_lang, target_lang)
        return service.translate_cached(text)

    def translate_document(
        self,
        input_path: str,
        source_lang: str,
        target_lang: str,
        tier: UserTier,
        preserve_original: bool = True
    ) -> str:
        """
        Translate document (PDF, Word, Excel, PowerPoint)

        Args:
            input_path: Path to input document
            source_lang: Source language code
            target_lang: Target language code
            tier: User subscription tier
            preserve_original: For PDFs, whether to preserve original text

        Returns:
            Path to translated document
        """
        # Get translation service based on tier
        translation_service = self._get_translation_service(tier, source_lang, target_lang)

        # Create document translator with the service
        # We need to manually construct the translator
        from document_translator import PDFTranslator, ExcelTranslator, WordTranslator, PowerPointTranslator

        try:
            import pymupdf
        except ImportError:
            pymupdf = None

        try:
            from openpyxl import load_workbook
        except ImportError:
            load_workbook = None

        try:
            from docx import Document
        except ImportError:
            Document = None

        try:
            from pptx import Presentation
        except ImportError:
            Presentation = None

        # Create custom translator instance
        translator = DocumentTranslator.__new__(DocumentTranslator)
        translator.translation_service = translation_service
        translator.pdf_translator = PDFTranslator(translation_service) if pymupdf else None
        translator.excel_translator = ExcelTranslator(translation_service) if load_workbook else None
        translator.word_translator = WordTranslator(translation_service) if Document else None
        translator.ppt_translator = PowerPointTranslator(translation_service) if Presentation else None
        translator.translate_document = DocumentTranslator.translate_document.__get__(translator, DocumentTranslator)
        translator._print_summary = DocumentTranslator._print_summary.__get__(translator, DocumentTranslator)

        # Generate output path
        input_file = Path(input_path)
        output_path = str(input_file.parent / f"{input_file.stem}_translated{input_file.suffix}")

        # Translate
        translator.translate_document(
            input_path=input_path,
            output_path=output_path,
            preserve_original=preserve_original
        )

        return output_path

    def translate_image(
        self,
        input_path: str,
        source_lang: str,
        target_lang: str,
        tier: UserTier
    ) -> str:
        """
        Translate image using OCR + translation + rendering

        Args:
            input_path: Path to input image
            source_lang: Source language code
            target_lang: Target language code
            tier: User subscription tier

        Returns:
            Path to translated image
        """
        if not IMAGE_TRANSLATOR_AVAILABLE:
            raise ImportError("Image translator not available. Please install image-translator dependencies.")

        # Get translation service based on tier
        translation_service = self._get_translation_service(tier, source_lang, target_lang)

        # Create image translation pipeline config
        config = {
            'ocr': {
                'languages': self._get_ocr_languages(source_lang),
                'use_gpu': False
            },
            'translation': {
                'engine': 'custom',
                'source_lang': source_lang,
                'target_lang': target_lang,
                'cache_translations': True
            },
            'inpainting': {
                'method': 'telea',
                'radius': 5
            },
            'fonts': {
                'chinese_simplified': '/System/Library/Fonts/PingFang.ttc',
                'chinese_traditional': '/System/Library/Fonts/PingFang.ttc',
                'english': '/System/Library/Fonts/Helvetica.ttc',
                'japanese': '/System/Library/Fonts/Hiragino Sans GB.ttc',
                'korean': '/System/Library/Fonts/AppleSDGothicNeo.ttc'
            },
            'rendering': {
                'default_font_size': 12,
                'auto_font_size': True
            }
        }

        # Create or get cached pipeline
        cache_key = f"{tier.value}_{source_lang}_{target_lang}"
        if cache_key not in self.image_pipeline_cache:
            pipeline = TranslationPipeline(config)
            # Inject our translation service
            pipeline.translator.translator = translation_service
            self.image_pipeline_cache[cache_key] = pipeline
        else:
            pipeline = self.image_pipeline_cache[cache_key]

        # Generate output path
        input_file = Path(input_path)
        output_path = str(input_file.parent / f"{input_file.stem}_translated{input_file.suffix}")

        # Translate image
        pipeline.translate_image(
            input_path=input_path,
            output_path=output_path,
            visualize=False
        )

        return output_path

    def _get_ocr_languages(self, lang_code: str) -> list:
        """Convert language code to OCR language list"""
        lang_map = {
            'en': ['en'],
            'zh-CN': ['ch', 'en'],
            'zh-TW': ['chinese_cht', 'en'],
            'ja': ['japan', 'en'],
            'ko': ['korean', 'en'],
            'ar': ['arabic', 'en'],
            'fr': ['french', 'en'],
            'de': ['german', 'en'],
            'es': ['spanish', 'en'],
            'pt': ['portuguese', 'en'],
            'ru': ['cyrillic', 'en'],
            'it': ['italian', 'en']
        }
        return lang_map.get(lang_code, ['en'])

    def get_tier_limits(self, tier: UserTier) -> Dict:
        """
        Get usage limits for a tier

        Returns:
            Dictionary with tier information and limits
        """
        limits = {
            UserTier.FREE: {
                'name': 'Free',
                'price': 0,
                'text_chars_per_month': 500000,  # 500K characters
                'documents_per_month': 3,
                'images_per_month': 10,
                'max_file_size_mb': 10,
                'features': [
                    'Google Translate',
                    'Basic text translation',
                    'Limited document translation',
                    'Limited image translation'
                ]
            },
            UserTier.DEEPL: {
                'name': 'DeepL Pro',
                'price': 9.99,
                'text_chars_per_month': 5000000,  # 5M characters
                'documents_per_month': 50,
                'images_per_month': 100,
                'max_file_size_mb': 50,
                'features': [
                    'DeepL API translation',
                    'Higher quality translations',
                    'Unlimited text translation',
                    'More document formats',
                    'Priority processing'
                ]
            },
            UserTier.AI: {
                'name': 'AI Premium',
                'price': 29.99,
                'text_chars_per_month': -1,  # Unlimited
                'documents_per_month': -1,  # Unlimited
                'images_per_month': -1,  # Unlimited
                'max_file_size_mb': 100,
                'features': [
                    'AI-powered translation (ChatGPT/Claude)',
                    'Best translation quality',
                    'Context-aware translation',
                    'Unlimited everything',
                    'Batch processing',
                    'API access',
                    'Priority support'
                ]
            }
        }

        return limits.get(tier, limits[UserTier.FREE])

    def get_tier_info(self) -> Dict:
        """Get information about all tiers"""
        return {
            'free': self.get_tier_limits(UserTier.FREE),
            'deepl': self.get_tier_limits(UserTier.DEEPL),
            'ai': self.get_tier_limits(UserTier.AI)
        }
